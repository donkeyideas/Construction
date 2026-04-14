#!/usr/bin/env python3
"""Shared XLSX builder and financial verification for mock data generators."""

import sys
import os
from collections import defaultdict


# Sheet order matches DEPENDENCY_ORDER in xlsx-parser.ts
# Sheet names must match SHEET_ENTITY_MAP (case-insensitive)
SHEET_CONFIG = [
    ("chart_of_accounts", "Chart of Accounts"),
    ("bank_accounts", "Bank Accounts"),
    ("properties", "Properties"),
    ("units", "Units"),
    ("projects", "Projects"),
    ("contacts", "Contacts"),
    ("vendors", "Vendors"),
    ("equipment", "Equipment"),
    ("phases", "Phases"),
    ("contracts", "Contracts"),
    ("opportunities", "Opportunities"),
    ("bids", "Bids"),
    ("leases", "Leases"),
    ("maintenance", "Maintenance"),
    ("budget_lines", "Budget Lines"),
    ("invoices", "Invoices"),
    ("journal_entries", "Journal Entries"),
    ("time_entries", "Time Entries"),
    ("change_orders", "Change Orders"),
    ("daily_logs", "Daily Logs"),
    ("rfis", "RFIs"),
    ("safety_incidents", "Safety Incidents"),
    ("safety_inspections", "Safety Inspections"),
    ("toolbox_talks", "Toolbox Talks"),
    ("equipment_assignments", "Equipment Assignments"),
    ("equipment_maintenance", "Equipment Maintenance"),
    ("submittals", "Submittals"),
    ("tasks", "Tasks"),
    ("property_expenses", "Property Expenses"),
    ("estimates", "Estimates"),
    ("certifications", "Certifications"),
]


def build_xlsx(all_sheets, output_path):
    """Build XLSX with proper sheet names for the Buildwrk import system."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_rows = 0
    for key, sheet_name in SHEET_CONFIG:
        if key not in all_sheets or not all_sheets[key]:
            continue
        data = all_sheets[key]
        ws = wb.create_sheet(title=sheet_name)

        headers = list(data[0].keys())
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)

        for r_idx, row in enumerate(data, 2):
            for c_idx, h in enumerate(headers, 1):
                val = row.get(h, "")
                ws.cell(row=r_idx, column=c_idx, value=str(val) if val else "")

        total_rows += len(data)
        print(f"  {sheet_name:<30} {len(data):>6} rows")

    wb.save(output_path)
    print(f"\nTotal rows: {total_rows}")
    print(f"Saved to: {output_path}")
    return output_path


def verify_financials(je_rows, invoices, target_ni=None, account_names=None):
    """Verify financial integrity of generated data."""
    print("\n" + "="*70)
    print("FINANCIAL VERIFICATION")
    print("="*70)

    if account_names is None:
        account_names = {}

    # 1. Verify all JEs balance individually
    je_totals = defaultdict(lambda: [0.0, 0.0])
    for r in je_rows:
        dr = float(r["debit"]) if r["debit"] else 0
        cr = float(r["credit"]) if r["credit"] else 0
        je_totals[r["entry_number"]][0] += dr
        je_totals[r["entry_number"]][1] += cr

    unbalanced = 0
    for je, (dr, cr) in je_totals.items():
        if abs(dr - cr) > 0.02:
            print(f"  UNBALANCED JE: {je} DR={dr:.2f} CR={cr:.2f}")
            unbalanced += 1

    if unbalanced == 0:
        print(f"[OK] All {len(je_totals)} journal entries balance")
    else:
        print(f"[FAIL] {unbalanced} unbalanced journal entries!")

    # 2. Build GL balances from JEs
    gl_balances = defaultdict(float)
    for r in je_rows:
        acct = int(r["account_number"])
        dr = float(r["debit"]) if r["debit"] else 0
        cr = float(r["credit"]) if r["credit"] else 0
        gl_balances[acct] += dr - cr

    # 3. Simulate invoice auto-JE impact
    for inv in invoices:
        amt = float(inv["amount"])
        ret = float(inv["retainage_held"]) if inv.get("retainage_held") else 0
        gl = int(inv["gl_account"]) if inv.get("gl_account") else None
        is_paid = inv["status"] == "paid"

        if inv["invoice_type"] == "receivable":
            net_ar = amt - ret
            gl_balances[1010] += net_ar
            if ret > 0:
                gl_balances[1020] += ret
            if gl:
                gl_balances[gl] -= amt
            if is_paid:
                gl_balances[1000] += net_ar
                gl_balances[1010] -= net_ar
        elif inv["invoice_type"] == "payable":
            net_ap = amt - ret
            if gl:
                gl_balances[gl] += amt
            gl_balances[2000] -= net_ap
            if ret > 0:
                gl_balances[2010] -= ret
            if is_paid:
                gl_balances[2000] += net_ap
                gl_balances[1000] -= net_ap

    # 4. Print trial balance
    print(f"\n{'TRIAL BALANCE (Estimated Post-Import)':^70}")
    print("-"*70)
    print(f"{'Account':>6}  {'Name':<42}  {'Debit':>12}  {'Credit':>12}")
    print("-"*70)

    total_dr = 0
    total_cr = 0
    for acct in sorted(gl_balances.keys()):
        bal = gl_balances[acct]
        if abs(bal) < 0.01:
            continue
        name = account_names.get(acct, str(acct))
        if bal > 0:
            print(f"  {acct:>4}  {name:<42}  {bal:>12,.2f}  {'':>12}")
            total_dr += bal
        else:
            print(f"  {acct:>4}  {name:<42}  {'':>12}  {-bal:>12,.2f}")
            total_cr += -bal

    print("-"*70)
    print(f"        {'TOTALS':<42}  {total_dr:>12,.2f}  {total_cr:>12,.2f}")
    diff = total_dr - total_cr
    print(f"        {'DIFFERENCE':<42}  {diff:>12,.2f}")

    if abs(diff) < 1.0:
        print("[OK] Trial balance is within $1.00")
    else:
        print(f"[WARN] Trial balance off by ${diff:,.2f}")

    # 5. Income Statement summary
    print(f"\n{'INCOME STATEMENT SUMMARY (Estimated)':^70}")
    print("-"*70)
    revenue = 0
    for acct in range(4000, 5000):
        if acct in gl_balances:
            revenue += -gl_balances[acct]
    expenses = 0
    for acct in range(5000, 8000):
        if acct in gl_balances:
            expenses += gl_balances[acct]

    net_income = revenue - expenses
    print(f"  Total Revenue:    ${revenue:>15,.2f}")
    print(f"  Total Expenses:   ${expenses:>15,.2f}")
    print(f"  Net Income:       ${net_income:>15,.2f}")

    if target_ni is not None:
        print(f"  Target NI:        ${target_ni:>15,.2f}")
        ni_diff = net_income - target_ni
        if abs(ni_diff) < 100000:
            print(f"  [OK] Net income within $100K of target (diff: ${ni_diff:,.2f})")
        else:
            print(f"  [WARN] Net income differs from target by ${ni_diff:,.2f}")

    return net_income
