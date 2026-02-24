#!/usr/bin/env python3
"""
Generate comprehensive mock data for Pinnacle Pacific Builders LLC.
Dual business: Airport construction ($480M) + 500-unit luxury condo ($185M + property mgmt).

Output: Pinnacle_Pacific_Builders_Import.xlsx

Usage: python generate_pinnacle_data.py
Requires: pip install openpyxl
"""

import sys
import os
from collections import defaultdict

# Add current dir to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from part01_constants import *
from part02_foundation import generate_chart_of_accounts, generate_bank_accounts, generate_properties, generate_units
from part03_master_data import generate_projects, generate_contacts, generate_vendors, generate_equipment
from part04_project_mgmt import generate_phases, generate_tasks, generate_contracts, generate_opportunities, generate_bids
from part05_field_ops import generate_daily_logs_airport, generate_daily_logs_condo, generate_rfis, generate_submittals, generate_change_orders
from part06_safety_labor import (generate_safety_incidents, generate_safety_inspections,
                                  generate_toolbox_talks, generate_certifications,
                                  generate_time_entries, generate_equipment_maintenance)
from part07_financials import generate_invoices, generate_journal_entries
from part08_property_mgmt import generate_leases, generate_maintenance, generate_property_expenses
from part09_assignments_estimates import generate_equipment_assignments, generate_estimates


def verify_financials(je_rows, invoices):
    """Verify financial integrity of the generated data."""
    print("\n" + "="*70)
    print("FINANCIAL VERIFICATION")
    print("="*70)

    # 1. Verify all JEs balance individually
    je_totals = defaultdict(lambda: [0.0, 0.0])
    for r in je_rows:
        dr = float(r["debit"]) if r["debit"] else 0
        cr = float(r["credit"]) if r["credit"] else 0
        je_totals[r["entry_number"]][0] += dr
        je_totals[r["entry_number"]][1] += cr

    unbalanced = 0
    for je, (dr, cr) in je_totals.items():
        if abs(dr - cr) > 0.02:
            print(f"  UNBALANCED JE: {je} DR={dr:.2f} CR={cr:.2f}")
            unbalanced += 1

    if unbalanced == 0:
        print(f"[OK] All {len(je_totals)} journal entries balance")
    else:
        print(f"[FAIL] {unbalanced} unbalanced journal entries!")

    # 2. Build trial balance from JEs only (non-AR/AP)
    gl_balances = defaultdict(float)
    for r in je_rows:
        acct = int(r["account_number"])
        dr = float(r["debit"]) if r["debit"] else 0
        cr = float(r["credit"]) if r["credit"] else 0
        gl_balances[acct] += dr - cr  # Debit-normal

    # 3. Estimate invoice auto-JE impact
    # Receivable: DR AR / CR gl_account (revenue or retained earnings)
    #             If paid: DR Cash / CR AR
    #             If retainage: DR Retainage Recv / CR AR (reduce AR by retainage)
    # Payable:   DR gl_account (expense or retained earnings) / CR AP
    #            If paid: DR AP / CR Cash
    #            If retainage: DR AP / CR Retainage Pay

    for inv in invoices:
        amt = float(inv["amount"])
        ret = float(inv["retainage_held"]) if inv.get("retainage_held") else 0
        gl = int(inv["gl_account"]) if inv.get("gl_account") else None
        is_paid = inv["status"] == "paid"

        if inv["invoice_type"] == "receivable":
            net_ar = amt - ret
            gl_balances[1010] += net_ar      # DR AR (net of retainage)
            if ret > 0:
                gl_balances[1020] += ret     # DR Retainage Receivable
            if gl:
                gl_balances[gl] -= amt       # CR Revenue/RE
            if is_paid:
                gl_balances[1000] += net_ar  # DR Cash
                gl_balances[1010] -= net_ar  # CR AR

        elif inv["invoice_type"] == "payable":
            net_ap = amt - ret
            if gl:
                gl_balances[gl] += amt       # DR Expense/RE
            gl_balances[2000] -= net_ap      # CR AP (net of retainage)
            if ret > 0:
                gl_balances[2010] -= ret     # CR Retainage Payable
            if is_paid:
                gl_balances[2000] += net_ap  # DR AP
                gl_balances[1000] -= net_ap  # CR Cash

    # 4. Print trial balance
    print(f"\n{'TRIAL BALANCE (Estimated Post-Import)':^70}")
    print("-"*70)
    print(f"{'Account':>6}  {'Name':<42}  {'Debit':>12}  {'Credit':>12}")
    print("-"*70)

    total_dr = 0
    total_cr = 0
    for acct in sorted(gl_balances.keys()):
        bal = gl_balances[acct]
        if abs(bal) < 0.01:
            continue
        # Get account name
        name = str(acct)
        for row_acct, (row_name, _) in list(DIRECT_COST_ACCOUNTS.items()) + list(PROPERTY_MGMT_ACCOUNTS.items()) + list(OVERHEAD_ACCOUNTS.items()):
            if row_acct == acct:
                name = row_name
                break
        acct_names = {
            1000: "Cash & Equivalents", 1010: "Accounts Receivable", 1020: "Retainage Receivable",
            1030: "Costs in Excess", 1040: "Prepaid Expenses", 1050: "Rent Receivable",
            1100: "Equipment & Vehicles", 1110: "Accum Dep - Equipment",
            1120: "Buildings & Improvements", 1130: "Accum Dep - Buildings",
            1200: "Land", 1300: "Security Deposits",
            2000: "Accounts Payable", 2010: "Retainage Payable", 2020: "Accrued Payroll",
            2030: "Accrued Expenses", 2040: "Billings in Excess", 2050: "Sales Tax Payable",
            2060: "Deferred Rental Revenue", 2100: "Equipment Financing",
            2200: "Construction LOC", 2210: "Mortgage Payable",
            3000: "Owners Capital", 3010: "Retained Earnings",
            4000: "Contract Revenue - Airport", 4010: "Contract Revenue - Condo",
            4100: "Rental Income", 4200: "Change Order Revenue",
            6100: "Depreciation - Equipment", 6110: "Depreciation - Buildings",
            7000: "Interest Expense",
        }
        if acct in acct_names:
            name = acct_names[acct]

        if bal > 0:
            print(f"  {acct:>4}  {name:<42}  {bal:>12,.2f}  {'':>12}")
            total_dr += bal
        else:
            print(f"  {acct:>4}  {name:<42}  {'':>12}  {-bal:>12,.2f}")
            total_cr += -bal

    print("-"*70)
    print(f"        {'TOTALS':<42}  {total_dr:>12,.2f}  {total_cr:>12,.2f}")
    diff = total_dr - total_cr
    print(f"        {'DIFFERENCE':<42}  {diff:>12,.2f}")

    if abs(diff) < 1.0:
        print("[OK] Trial balance is within $1.00")
    else:
        print(f"[WARN] Trial balance off by ${diff:,.2f}")

    # 5. Income Statement summary
    print(f"\n{'INCOME STATEMENT SUMMARY (Estimated)':^70}")
    print("-"*70)
    revenue = 0
    for acct in range(4000, 5000):
        if acct in gl_balances:
            revenue += -gl_balances[acct]  # Revenue is credit (negative in debit-normal)
    expenses = 0
    for acct in range(5000, 8000):
        if acct in gl_balances:
            expenses += gl_balances[acct]  # Expenses are debit (positive)

    net_income = revenue - expenses
    print(f"  Total Revenue:    ${revenue:>15,.2f}")
    print(f"  Total Expenses:   ${expenses:>15,.2f}")
    print(f"  Net Income:       ${net_income:>15,.2f}")
    print(f"  Target NI:        ${NET_INCOME:>15,.2f}")
    ni_diff = net_income - NET_INCOME
    if abs(ni_diff) < 100000:
        print(f"  [OK] Net income within $100K of target (diff: ${ni_diff:,.2f})")
    else:
        print(f"  [WARN] Net income differs from target by ${ni_diff:,.2f}")


def build_xlsx(all_sheets):
    """Build XLSX with proper sheet names for the Buildwrk import system."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet order matches DEPENDENCY_ORDER in xlsx-parser.ts
    # Sheet names must match SHEET_ENTITY_MAP (case-insensitive, spaces)
    SHEET_CONFIG = [
        ("chart_of_accounts", "Chart of Accounts"),
        ("bank_accounts", "Bank Accounts"),
        ("properties", "Properties"),
        ("units", "Units"),
        ("projects", "Projects"),
        ("contacts", "Contacts"),
        ("vendors", "Vendors"),
        ("equipment", "Equipment"),
        ("phases", "Phases"),
        ("contracts", "Contracts"),
        ("opportunities", "Opportunities"),
        ("bids", "Bids"),
        ("leases", "Leases"),
        ("maintenance", "Maintenance"),
        ("invoices", "Invoices"),
        ("journal_entries", "Journal Entries"),
        ("time_entries", "Time Entries"),
        ("change_orders", "Change Orders"),
        ("daily_logs", "Daily Logs"),
        ("rfis", "RFIs"),
        ("safety_incidents", "Safety Incidents"),
        ("safety_inspections", "Safety Inspections"),
        ("toolbox_talks", "Toolbox Talks"),
        ("equipment_assignments", "Equipment Assignments"),
        ("equipment_maintenance", "Equipment Maintenance"),
        ("submittals", "Submittals"),
        ("tasks", "Tasks"),
        ("property_expenses", "Property Expenses"),
        ("estimates", "Estimates"),
        ("certifications", "Certifications"),
    ]

    total_rows = 0
    for key, sheet_name in SHEET_CONFIG:
        if key not in all_sheets or not all_sheets[key]:
            continue
        data = all_sheets[key]
        ws = wb.create_sheet(title=sheet_name)

        headers = list(data[0].keys())
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)

        for r_idx, row in enumerate(data, 2):
            for c_idx, h in enumerate(headers, 1):
                val = row.get(h, "")
                ws.cell(row=r_idx, column=c_idx, value=str(val) if val else "")

        total_rows += len(data)
        print(f"  {sheet_name:<30} {len(data):>6} rows")

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "Pinnacle_Pacific_Builders_Import.xlsx")
    wb.save(out_path)
    print(f"\nTotal rows: {total_rows}")
    print(f"Saved to: {out_path}")
    return out_path


def main():
    print("="*70)
    print("PINNACLE PACIFIC BUILDERS - MOCK DATA GENERATOR")
    print("="*70)
    print(f"Airport: {AIRPORT['name']} (${AIRPORT['budget']:,.0f})")
    print(f"Condo:   {CONDO['name']} (${CONDO['budget']:,.0f})")
    print(f"Target Net Income: ${NET_INCOME:,.0f}")
    print()

    # Generate all data
    print("Generating data...")
    all_sheets = {}

    all_sheets["chart_of_accounts"] = generate_chart_of_accounts()
    all_sheets["bank_accounts"] = generate_bank_accounts()
    all_sheets["properties"] = generate_properties()

    units = generate_units()
    all_sheets["units"] = units

    all_sheets["projects"] = generate_projects()
    all_sheets["contacts"] = generate_contacts()
    all_sheets["vendors"] = generate_vendors()
    all_sheets["equipment"] = generate_equipment()
    all_sheets["phases"] = generate_phases()
    all_sheets["tasks"] = generate_tasks()
    all_sheets["contracts"] = generate_contracts()
    all_sheets["opportunities"] = generate_opportunities()
    all_sheets["bids"] = generate_bids()

    all_sheets["daily_logs"] = generate_daily_logs_airport() + generate_daily_logs_condo()
    all_sheets["rfis"] = generate_rfis()
    all_sheets["submittals"] = generate_submittals()
    all_sheets["change_orders"] = generate_change_orders()

    all_sheets["safety_incidents"] = generate_safety_incidents()
    all_sheets["safety_inspections"] = generate_safety_inspections()
    all_sheets["toolbox_talks"] = generate_toolbox_talks()
    all_sheets["certifications"] = generate_certifications()
    all_sheets["time_entries"] = generate_time_entries()
    all_sheets["equipment_maintenance"] = generate_equipment_maintenance()

    all_sheets["equipment_assignments"] = generate_equipment_assignments()
    all_sheets["estimates"] = generate_estimates()

    all_sheets["invoices"] = generate_invoices()
    all_sheets["journal_entries"] = generate_journal_entries()

    all_sheets["leases"] = generate_leases(units)
    all_sheets["maintenance"] = generate_maintenance()
    all_sheets["property_expenses"] = generate_property_expenses()

    # Print summary
    print(f"\nData generation complete:")
    for key, data in all_sheets.items():
        print(f"  {key:<30} {len(data):>6} rows")

    # Verify financials
    verify_financials(all_sheets["journal_entries"], all_sheets["invoices"])

    # Build XLSX
    print(f"\n{'BUILDING XLSX':^70}")
    print("-"*70)
    out_path = build_xlsx(all_sheets)

    print(f"\n{'='*70}")
    print("DONE! Import the file via Buildwrk Settings > Import Data")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
